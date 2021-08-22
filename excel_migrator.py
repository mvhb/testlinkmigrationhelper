from openpyxl import load_workbook
import xml.etree.ElementTree as ET
import xlsxwriter
import re
import os


def create_excel_name(xml_name):
    return xml_name[:-4] + '_to_excel.xlsx'


def get_excel_path(excel_name):
    excel_path = os.path.abspath(".") + "/excel/" + excel_name
    return excel_path


def creating_excel_file_with_headers(excel_path):
    workbook = xlsxwriter.Workbook(excel_path)
    worksheet = workbook.add_worksheet()

    # Columns header
    worksheet.write('A1', 'TESTCASE NAME')
    worksheet.write('B1', 'NODE ORDER')
    worksheet.write('C1', 'EXTERNAL ID')
    worksheet.write('D1', 'VERSION')
    worksheet.write('E1', 'SUMMARY')
    worksheet.write('F1', 'PRECONDITIONS')
    worksheet.write('G1', 'EXECUTION TYPE')
    worksheet.write('H1', 'IMPORTANCE')
    worksheet.write('I1', 'KEYWORDS')

    return workbook, worksheet


def read_xml_and_populate_excel(workbook, worksheet, xml_name):
    count = 2  # Header is the first line

    # Reading the XML file
    tree = ET.parse(os.path.abspath(".") + "/xml/" + xml_name)
    root = tree.getroot()

    for testcase in root.findall('testcase'):

        testcase_title = testcase.get('name')
        worksheet.write('A' + str(count), testcase_title)

        node_order = testcase.find('node_order').text
        worksheet.write('B' + str(count), node_order)

        external_id = testcase.find('externalid').text
        worksheet.write('C' + str(count), external_id)

        version = testcase.find('version').text
        worksheet.write('D' + str(count), version)

        summary = testcase.find('summary').text
        worksheet.write('E' + str(count), summary)

        pre_condition = testcase.find('preconditions').text
        worksheet.write('F' + str(count), pre_condition)

        execution_type = testcase.find('execution_type').text
        worksheet.write('G' + str(count), execution_type)

        importance = testcase.find('importance').text
        worksheet.write('H' + str(count), importance)

        keywords = ""
        for keyword in testcase.find('keywords'):
            keywords += keyword.get('name') + ", "
        worksheet.write('I' + str(count), keywords[:-2])

        count += 1

    workbook.close()


def remove_html_tags(excel_path):
    book = load_workbook(excel_path)
    sheet = book.active
    summary_column = sheet['E'] # SUMMARY column

    cleanr = re.compile('<.*?>')

    for row in summary_column:
        if row.value != None:
            row.value = re.sub(cleanr, '', row.value)
            row.value = row.value.replace("&lt;", "<")
            row.value = row.value.replace("&gt;", ">")
            row.value = row.value.replace("-&gt;", ">")
            row.value = row.value.replace("&nbsp;", " ")
            row.value = row.value.replace("&quot;", '"')
    book.save(excel_path)


def translate_automation_status(excel_path):
    book = load_workbook(excel_path)
    sheet = book.active
    column_automation_status = sheet['G']

    for row in column_automation_status:
        if row.value == "1":
            row.value = "Manual"
        elif row.value == "2":
            row.value = "Automated"

    book.save(excel_path)


if __name__ == "__main__":
    while True:
        xml_name = input("Type the name of the XML file with .xml: ")
        if xml_name.upper() != "QUIT":
            excel_name = create_excel_name(xml_name)
            excel_path = get_excel_path(excel_name)
            workbook, worksheet = creating_excel_file_with_headers(excel_path)
            read_xml_and_populate_excel(workbook, worksheet, xml_name)
            remove_html_tags(excel_path)
            translate_automation_status(excel_path)
            print("The excel '{excel_name}' was created on '{excel_path}'".format(
                excel_name=excel_name, excel_path=excel_path))
        else:
            break
